#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
CONVERSOR DE NOTAS - JSON antiguo → Formato nuevo del sistema
Convierte notas de sistemita_datos.json al formato Calificativo
"""

import json
from datetime import datetime
from collections import defaultdict
import os

# ────────────────────────────────────────────────────────────────────────────
# PASO 1: CARGAR DATOS
# ────────────────────────────────────────────────────────────────────────────

def cargar_json(ruta):
    """Carga el archivo JSON antiguo"""
    try:
        with open(ruta, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"❌ Error cargando JSON: {e}")
        return None

# ────────────────────────────────────────────────────────────────────────────
# PASO 2: MAPEAR FORMATOS
# ────────────────────────────────────────────────────────────────────────────

def convertir_nota_numerica_a_calificativo(valor):
    """Convierte 0-20 a C, B, A, AD"""
    valor = float(valor) if isinstance(valor, (int, str)) else 0
    if valor >= 18: return 'AD'
    if valor >= 15: return 'A'
    if valor >= 11: return 'B'
    return 'C'

def contar_adecuadas(calificacion_dict):
    """Cuenta respuestas adecuadas en un dict de calificaciones"""
    if not isinstance(calificacion_dict, dict):
        return 0, 0, []

    adecuadas = sum(1 for v in calificacion_dict.values() if v == 'ADECUADA')
    total = len(calificacion_dict)
    claves = list(calificacion_dict.keys())
    marcados = [v == 'ADECUADA' for v in calificacion_dict.values()]

    return adecuadas, total, claves, marcados

def calcular_calificativo(adecuadas, total):
    """Calcula calificativo basado en porcentaje de adecuadas"""
    if total == 0:
        return 'C'
    porcentaje = (adecuadas / total) * 100
    if porcentaje >= 80: return 'AD'
    if porcentaje >= 60: return 'A'
    if porcentaje >= 40: return 'B'
    return 'C'

# ────────────────────────────────────────────────────────────────────────────
# PASO 3: CONVERTIR NOTAS
# ────────────────────────────────────────────────────────────────────────────

def procesar_notas(datos_json):
    """Procesa y convierte las notas al nuevo formato"""

    notas_convertidas = []
    notas_sin_mapeo = []

    if 'notas' not in datos_json:
        print("❌ No se encontró 'notas' en el JSON")
        return [], []

    for nota_vieja in datos_json['notas']:
        calificacion_raw = nota_vieja.get('calificacion', '')

        try:
            # CASO 1: Calificación es número (examen numérico)
            if isinstance(calificacion_raw, str) and calificacion_raw.isdigit():
                calificativo = convertir_nota_numerica_a_calificativo(int(calificacion_raw))
                nota_nueva = {
                    'alumnoId': nota_vieja.get('student_id', ''),
                    'competenciaId': nota_vieja.get('competencia', ''),
                    'columnaId': nota_vieja.get('instrumento', ''),
                    'instrumento': nota_vieja.get('instrumento', ''),
                    'columnaIds': [nota_vieja.get('instrumento', '')],
                    'marcados': [],
                    'claves': [],
                    'notaNumerica': int(calificacion_raw),
                    'calificativo': calificativo,
                    'esAD': calificativo == 'AD',
                    'fecha': nota_vieja.get('fecha', datetime.now().isoformat()),
                    'periodo': nota_vieja.get('periodo', 'Abril 2026'),
                    'alumnoNombre': nota_vieja.get('student_name', ''),
                    'competencia': nota_vieja.get('competencia', '')
                }

            # CASO 2: Calificación es diccionario (rúbrica con respuestas)
            elif isinstance(calificacion_raw, dict):
                adecuadas, total, claves, marcados = contar_adecuadas(calificacion_raw)
                calificativo = calcular_calificativo(adecuadas, total)

                nota_nueva = {
                    'alumnoId': nota_vieja.get('student_id', ''),
                    'competenciaId': nota_vieja.get('competencia', ''),
                    'columnaId': nota_vieja.get('instrumento', ''),
                    'instrumento': nota_vieja.get('instrumento', ''),
                    'columnaIds': [nota_vieja.get('instrumento', '')],
                    'marcados': marcados,
                    'claves': claves,
                    'respuestasAdecuadas': adecuadas,
                    'respuestasTotal': total,
                    'calificativo': calificativo,
                    'esAD': calificativo == 'AD',
                    'fecha': nota_vieja.get('fecha', datetime.now().isoformat()),
                    'periodo': nota_vieja.get('periodo', 'Abril 2026'),
                    'alumnoNombre': nota_vieja.get('student_name', ''),
                    'competencia': nota_vieja.get('competencia', '')
                }
            else:
                notas_sin_mapeo.append(nota_vieja)
                continue

            notas_convertidas.append(nota_nueva)

        except Exception as e:
            print(f"⚠️ Error procesando nota {nota_vieja.get('id')}: {e}")
            notas_sin_mapeo.append(nota_vieja)

    return notas_convertidas, notas_sin_mapeo

# ────────────────────────────────────────────────────────────────────────────
# PASO 4: GENERAR JSON PARA CARGAR
# ────────────────────────────────────────────────────────────────────────────

def guardar_json_convertido(notas, ruta_salida):
    """Guarda las notas convertidas en JSON"""
    with open(ruta_salida, 'w', encoding='utf-8') as f:
        json.dump(notas, f, indent=2, ensure_ascii=False)
    print(f"✅ JSON guardado: {ruta_salida}")
    return ruta_salida

# ────────────────────────────────────────────────────────────────────────────
# PASO 5: GENERAR EXCEL CON RESUMEN
# ────────────────────────────────────────────────────────────────────────────

def generar_excel_resumen(notas_convertidas, datos_json, ruta_excel):
    """Genera Excel con notas por alumno, unidad, grado"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("⚠️ openpyxl no instalado. Instalando...")
        os.system('pip install openpyxl')
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notas por Alumno"

    # Headers
    headers = ['Alumno', 'Grado', 'Competencia', 'Instrumento', 'Respuestas Adecuadas', 'Calificativo', 'Fecha', 'Periodo']
    ws.append(headers)

    # Estilos
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Datos
    for nota in notas_convertidas:
        respuestas = f"{nota.get('respuestasAdecuadas', '')}/{nota.get('respuestasTotal', '')}"
        ws.append([
            nota.get('alumnoNombre', ''),
            datos_json.get('estudiantes', [{}])[0].get('grado', ''),
            nota.get('competencia', ''),
            nota.get('instrumento', ''),
            respuestas if nota.get('respuestasAdecuadas') else '',
            nota.get('calificativo', ''),
            nota.get('fecha', ''),
            nota.get('periodo', '')
        ])

    # Ajustar ancho
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15

    wb.save(ruta_excel)
    print(f"✅ Excel guardado: {ruta_excel}")

# ────────────────────────────────────────────────────────────────────────────
# PASO 6: GENERAR EXCEL POR INSTRUMENTO
# ────────────────────────────────────────────────────────────────────────────

def generar_excel_por_instrumento(notas_convertidas, ruta_excel):
    """Genera Excel con detalle por instrumento y respuestas"""
    try:
        import openpyxl
    except ImportError:
        os.system('pip install openpyxl')
        import openpyxl

    # Agrupar por instrumento
    por_instrumento = defaultdict(list)
    for nota in notas_convertidas:
        instrumento = nota.get('instrumento', 'sin_instrumento')
        por_instrumento[instrumento].append(nota)

    wb = openpyxl.Workbook()

    # Hoja resumen
    ws_resumen = wb.active
    ws_resumen.title = "Resumen Instrumentos"
    ws_resumen['A1'] = "Instrumento"
    ws_resumen['B1'] = "Notas"

    row = 2
    for instr, notas in por_instrumento.items():
        ws_resumen[f'A{row}'] = instr
        ws_resumen[f'B{row}'] = len(notas)
        row += 1

    # Hoja detalle por instrumento
    for instrumento, notas in por_instrumento.items():
        ws = wb.create_sheet(title=instrumento[:31])  # Excel máx 31 chars

        headers = ['Alumno', 'Competencia', 'Respuestas', 'Calificativo', 'Detalle']
        ws.append(headers)

        for nota in notas:
            respuestas = ''
            if nota.get('claves'):
                detalle = ', '.join([f"{k}:{('✓' if m else '✗')}"
                                   for k, m in zip(nota['claves'], nota.get('marcados', []))])
            else:
                detalle = f"Nota numérica: {nota.get('notaNumerica', '')}/20"

            ws.append([
                nota.get('alumnoNombre', ''),
                nota.get('competencia', ''),
                f"{nota.get('respuestasAdecuadas', '')}/{nota.get('respuestasTotal', '')}" if nota.get('respuestasAdecuadas') else '',
                nota.get('calificativo', ''),
                detalle
            ])

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40

    wb.save(ruta_excel)
    print(f"✅ Excel por instrumento guardado: {ruta_excel}")

# ────────────────────────────────────────────────────────────────────────────
# MAIN
# ────────────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print("🔄 CONVERSOR DE NOTAS - Iniciando...\n")

    # Rutas
    ruta_json_antiguo = 'sistemita_datos.json'
    ruta_json_nuevo = 'notas_convertidas.json'
    ruta_excel_resumen = 'Notas_Resumen.xlsx'
    ruta_excel_instrumentos = 'Notas_Por_Instrumento.xlsx'

    # 1. Cargar
    print("1️⃣  Cargando datos...")
    datos = cargar_json(ruta_json_antiguo)
    if not datos:
        exit(1)

    # 2. Convertir
    print("2️⃣  Convirtiendo notas...")
    notas_nuevas, sin_mapeo = procesar_notas(datos)
    print(f"   ✅ {len(notas_nuevas)} notas convertidas")
    print(f"   ⚠️  {len(sin_mapeo)} notas sin mapeo")

    # 3. Guardar JSON
    print("3️⃣  Guardando JSON convertido...")
    guardar_json_convertido(notas_nuevas, ruta_json_nuevo)

    # 4. Generar Excel resumen
    print("4️⃣  Generando Excel con resumen...")
    generar_excel_resumen(notas_nuevas, datos, ruta_excel_resumen)

    # 5. Generar Excel por instrumento
    print("5️⃣  Generando Excel por instrumento...")
    generar_excel_por_instrumento(notas_nuevas, ruta_excel_instrumentos)

    # Resumen
    print("\n" + "="*60)
    print("✅ CONVERSIÓN COMPLETADA")
    print("="*60)
    print(f"📊 Notas convertidas: {len(notas_nuevas)}")
    print(f"📁 JSON nuevo: {ruta_json_nuevo}")
    print(f"📊 Excel resumen: {ruta_excel_resumen}")
    print(f"📊 Excel instrumentos: {ruta_excel_instrumentos}")
    print("\n⚠️  PRÓXIMO PASO:")
    print(f"1. Abre {ruta_json_nuevo}")
    print("2. Copia el contenido en la consola del navegador:")
    print("   localStorage.setItem('ie_calificativos_v2', JSON.stringify(datos))")
    print("="*60)
